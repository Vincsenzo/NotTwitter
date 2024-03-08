import openpyxl as xl
from datetime import datetime, timedelta

from .models import Posts


# Tutorial for excel file response: https://djangotricks.blogspot.com/2019/02/how-to-export-data-to-xlsx-files.html
def migrate_data_to_excel(response):
    posts = Posts.objects.all()
    wb = xl.Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value='Name of user')
    ws.cell(row=1, column=2, value='Post text')
    ws.cell(row=1, column=3, value='Time post created')

    for row_num, item in enumerate(posts, start=2):
        ws.cell(row=row_num, column=1, value=str(item.user))
        ws.cell(row=row_num, column=2, value=item.post_text)
        
        # Adjust time zone beacous django does not do it itself by default, see: https://stackoverflow.com/questions/16037020/djangos-timezone-now-does-not-show-the-right-time
        time_zone_bp = item.created + timedelta(hours=1)
        formatted_time = time_zone_bp.strftime('%Y-%m-%d %H:%M:%S')
        ws.cell(row=row_num, column=3, value=formatted_time)

    saved_file = wb.save(response)
    return saved_file
